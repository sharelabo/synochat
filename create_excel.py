import json
import pandas as pd
from datetime import datetime, date
import re
import os
from zoneinfo import ZoneInfo
import requests
from dotenv import load_dotenv
import openpyxl.styles

# 環境変数の読み込み
load_dotenv()

# 環境変数から設定を読み込む
DATA_DIR = os.getenv("DATA_DIR", "data")  # デフォルト値は "data"
TIMEZONE = os.getenv("TIMEZONE", "Asia/Tokyo")
WEBDAV_URL = os.getenv("WEBDAV_URL")
WEBDAV_USERNAME = os.getenv("WEBDAV_USERNAME")
WEBDAV_PASSWORD = os.getenv("WEBDAV_PASSWORD")
WEBDAV_FOLDER = os.getenv("WEBDAV_FOLDER", "/ExcelReports")  # WebDAV上の保存先フォルダ

# OpenAI API設定
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
OPENAI_API_BASE = "https://api.openai.com/v1"

# スクリプトのディレクトリを取得
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# DATA_DIRの設定
# Dockerコンテナ内の場合は /app/data を使用
if os.path.exists("/app"):
    DATA_DIR = "/app/data"
else:
    # 開発環境の場合は相対パスを使用
    DATA_DIR = os.path.join(SCRIPT_DIR, DATA_DIR)

print(f"使用するデータディレクトリ: {DATA_DIR}")


def get_period_start_end(dt):
    """日付から期間の開始日と終了日を取得"""
    current_date = dt.date()
    if current_date.day <= 10:
        # 前月11日から今月10日
        if current_date.month == 1:
            start_date = date(current_date.year - 1, 12, 11)
        else:
            start_date = date(current_date.year, current_date.month - 1, 11)
        end_date = date(current_date.year, current_date.month, 10)
    else:
        # 今月11日から来月10日
        if current_date.month == 12:
            end_date = date(current_date.year + 1, 1, 10)
        else:
            end_date = date(current_date.year, current_date.month + 1, 10)
        start_date = date(current_date.year, current_date.month, 11)
    return start_date, end_date


def get_period_filename(start_date, end_date):
    """期間に基づいてファイル名を生成"""
    return (
        f"messages_{start_date.strftime('%Y%m')}_{start_date.day:02d}-"
        f"{end_date.strftime('%Y%m')}_{end_date.day:02d}.json"
    )


def get_excel_filename(start_date, end_date):
    """期間に基づいてExcelファイル名を生成"""
    return (
        f"messages_{start_date.strftime('%Y%m')}_{start_date.day:02d}-"
        f"{end_date.strftime('%Y%m')}_{end_date.day:02d}.xlsx"
    )


def extract_tags(text):
    # タグを抽出（#タグ の形式）
    tags = re.findall(r"#\w+", text)  # '#'を含めてタグを抽出
    # タグを除いた本文を取得
    clean_text = re.sub(r"#\w+", "", text).strip()
    return tags, clean_text


def analyze_message_intent(message: str) -> str:
    """GPT-4o-mini-2024-07-18を使用してメッセージの意図を分析"""
    if not OPENAI_API_KEY:
        print("OPENAI_API_KEYが設定されていません")
        return "unknown"

    # 空のメッセージをチェック
    if not message or message.isspace():
        print("メッセージが空です")
        return "unknown"

    # #社長予定タグがある場合は出勤として判定
    if "#社長予定" in message:
        print("社長予定タグを検出: startとして判定")
        return "start"

    try:
        print(f"メッセージを分析中: {message}")

        # APIリクエストの設定
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {OPENAI_API_KEY}",
        }

        # プロンプトの設定
        prompt = f"""
以下のメッセージから、業務の開始・終了の意図を判定してください。
単純なキーワードマッチではなく、文脈や表現全体から判断してください。

判定は以下のいずれかの文字列で返してください：
- start: 業務開始のメッセージ（例：「業務開始します」「在宅勤務を開始します」）
- end: 業務終了のメッセージ（例：「業務終了します」「本日の業務を終えます」）
- other: それ以外のメッセージ

メッセージ: {message}

判定の際は以下の点を考慮してください：
1. 明示的な業務開始/終了の表現
   - 業務開始：「業務開始」「始業」「出勤」など、明確に業務の開始を宣言する表現
   - 業務終了：「業務終了」「退勤」「帰宅」など、明確に業務の終了を宣言する表現

2. 暗示的な業務開始/終了の表現
   - 業務開始：「おはようございます」＋その日の業務予定の列挙
   - 業務終了：「お疲れ様です」＋業務の完了報告

3. 文脈による判断
   - 業務開始：一日の業務予定を箇条書きで列挙している
   - 業務終了：一日の業務報告や総括をしている

4. 以下は業務開始/終了とみなさない
   - 単なる外出や戻り時間の報告
   - 書類や作業の依頼
   - 通常の業務連絡
   - 移動の報告
   - 作業状況の中間報告

回答は start, end, other のいずれかのみを返してください。
"""

        # APIリクエストを送信
        request_data = {
            "model": "gpt-4o-mini-2024-07-18",
            "messages": [
                {
                    "role": "system",
                    "content": "あなたはメッセージの意図を判定するアシスタントです。",
                },
                {"role": "user", "content": prompt},
            ],
            "temperature": 0.1,
            "max_tokens": 50,  # トークン数を増やす
        }
        print("APIリクエストを送信中...")
        print(f"リクエストURL: {OPENAI_API_BASE}/chat/completions")
        print(
            f"リクエストデータ: {json.dumps(request_data, ensure_ascii=False, indent=2)}"
        )

        response = requests.post(
            f"{OPENAI_API_BASE}/chat/completions",
            headers=headers,
            json=request_data,
        )

        # レスポンスのステータスコードを確認
        print(f"APIレスポンスステータス: {response.status_code}")
        if response.status_code != 200:
            print(f"APIエラーレスポンス: {response.text}")
            return "unknown"

        result = response.json()
        print(f"APIレスポンス: {json.dumps(result, ensure_ascii=False, indent=2)}")

        intent = result["choices"][0]["message"]["content"].strip().lower()
        print(f"判定結果: {intent}")

        # 有効な判定結果かチェック
        if intent not in ["start", "end", "other"]:
            print(f"不正な判定結果: {intent}")
            return "unknown"

        return intent

    except requests.exceptions.RequestException as e:
        print(f"APIリクエスト中にエラーが発生: {str(e)}")
        return "unknown"
    except json.JSONDecodeError as e:
        print(f"JSONデコードエラー: {str(e)}")
        print(f"レスポンス内容: {response.text}")
        return "unknown"
    except Exception as e:
        print(f"予期せぬエラーが発生: {str(e)}")
        import traceback

        print(f"詳細なエラー情報: {traceback.format_exc()}")
        return "unknown"


def classify_time(text, time_str):
    """メッセージの内容に基づいて時刻を分類"""
    intent = analyze_message_intent(text)

    if intent == "start":
        return time_str, "", ""
    elif intent == "end":
        return "", time_str, ""
    else:
        return "", "", time_str


def get_weekday_jp(dt):
    """日付から日本語の曜日を取得"""
    weekdays = ["月", "火", "水", "木", "金", "土", "日"]
    return f"{weekdays[dt.weekday()]}曜日"


def process_messages():
    """メッセージを処理してExcelファイルを生成"""
    # dataディレクトリ内の全てのJSONファイルを処理
    if not os.path.exists(DATA_DIR):
        print(f"ディレクトリが見つかりません: {DATA_DIR}")
        return

    json_files = [
        f
        for f in os.listdir(DATA_DIR)
        if f.startswith("messages_") and f.endswith(".json")
    ]

    if not json_files:
        print("処理対象のJSONファイルが見つかりません")
        return

    for json_file in json_files:
        json_path = os.path.join(DATA_DIR, json_file)
        print(f"JSONファイルを処理中: {json_file}")

        try:
            # JSONファイルを読み込む
            with open(json_path, "r", encoding="utf-8") as f:
                messages = json.load(f)

            # メッセージを処理
            processed_messages = {}
            for msg in messages:
                # タグと本文を分離
                text = msg.get(
                    "text", msg.get("message", "")
                )  # textフィールドがない場合はmessageフィールドを使用
                tags, clean_text = extract_tags(text)
                tags_str = (
                    ", ".join(tags) if tags else ""
                )  # '#'を含めたタグをそのまま使用

                # 時刻を分類
                received_at = msg.get("received_at", "")
                if received_at:
                    timestamp = datetime.fromisoformat(received_at)
                    time_str = timestamp.strftime("%H:%M")
                    start_time, end_time, unknown_time = classify_time(
                        clean_text, time_str
                    )

                    # ユーザー名でグループ化
                    username = msg.get("username", "未設定")
                    if username not in processed_messages:
                        processed_messages[username] = []

                    # データを追加
                    processed_messages[username].append(
                        {
                            "月": timestamp.strftime("%m"),
                            "日": timestamp.strftime("%d"),
                            "曜": get_weekday_jp(timestamp),
                            "出勤時刻": start_time,
                            "退社時刻": end_time,
                            "不明": unknown_time,
                            "タグ": tags_str,
                            "本文": clean_text,
                        }
                    )

            # Excelファイルとして保存
            excel_filename = os.path.splitext(json_file)[0] + ".xlsx"
            excel_path = os.path.join(DATA_DIR, excel_filename)

            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                # 各ユーザーのデータをシートとして保存
                for username, user_messages in processed_messages.items():
                    # DataFrameを作成し、列の順序を指定
                    df = pd.DataFrame(user_messages)
                    df = df[
                        [
                            "月",
                            "日",
                            "曜",
                            "出勤時刻",
                            "退社時刻",
                            "不明",
                            "タグ",
                            "本文",
                        ]
                    ]
                    # シート名としてユーザー名を使用（31文字以内）
                    sheet_name = username[:31]
                    # シートとして保存（ヘッダーの太字を無効化）
                    df.to_excel(
                        writer,
                        index=False,
                        sheet_name=sheet_name,
                    )

                    # ワークシートを取得
                    worksheet = writer.sheets[sheet_name]

                    # ヘッダーの太字を解除
                    for cell in worksheet[1]:
                        cell.font = openpyxl.styles.Font(bold=False)

                    # 罫線を削除
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.border = None

                    # 月と日の列を数値書式に設定
                    for row in worksheet.iter_rows(min_row=2):  # ヘッダー行をスキップ
                        # A列（月）の書式設定
                        row[0].number_format = "0"
                        # B列（日）の書式設定
                        row[1].number_format = "0"

                    # 列幅の自動調整
                    for idx, col in enumerate(df.columns):
                        # 列の最大文字数を計算（列名と内容の両方を考慮）
                        max_length = max(
                            df[col].astype(str).apply(len).max(),  # 内容の最大長
                            len(str(col)),  # 列名の長さ
                        )
                        # 文字幅から列幅を計算（1文字あたり1.2を掛けて余裕を持たせる）
                        adjusted_width = max_length * 1.2
                        # 列幅を設定（最小幅8、最大幅50）
                        worksheet.column_dimensions[chr(65 + idx)].width = min(
                            max(8, adjusted_width), 50
                        )

            print(f"Excelファイルを作成しました: {excel_path}")

        except Exception as e:
            print(f"ファイル処理中にエラーが発生しました: {json_file}")
            print(f"エラー: {str(e)}")
            continue


def upload_to_webdav(file_path, start_date, end_date):
    """WebDAVを使用してファイルをアップロード"""
    if not all([WEBDAV_URL, WEBDAV_USERNAME, WEBDAV_PASSWORD]):
        print("WebDAVの設定が不完全です。環境変数を確認してください。")
        return False

    try:
        # アップロード先のURLを構築
        filename = os.path.basename(file_path)
        encoded_folder = "/".join(
            requests.utils.quote(part) for part in WEBDAV_FOLDER.split("/")
        )
        encoded_filename = requests.utils.quote(filename)

        base_url = WEBDAV_URL.rstrip("/")
        upload_url = f"{base_url}/{encoded_folder}/{encoded_filename}"
        folder_url = f"{base_url}/{encoded_folder}/"

        print(f"アップロード先URL: {upload_url}")

        # フォルダの存在確認とMKCOL（作成）
        response = requests.request(
            "MKCOL", folder_url, auth=(WEBDAV_USERNAME, WEBDAV_PASSWORD), verify=False
        )
        print(f"フォルダ作成結果: {response.status_code}")

        # ファイルをアップロード
        with open(file_path, "rb") as f:
            response = requests.put(
                upload_url,
                data=f,
                auth=(WEBDAV_USERNAME, WEBDAV_PASSWORD),
                verify=False,
                headers={
                    "Content-Type": (
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet"
                    )
                },
            )

        if response.status_code in [200, 201, 204]:
            period_str = (
                f"{start_date.strftime('%Y/%m/%d')}～"
                f"{end_date.strftime('%Y/%m/%d')}"
            )
            print(f"ファイルのアップロードに成功しました: {filename}")
            print(f"保存先: {WEBDAV_FOLDER}")
            print(f"期間: {period_str}")
            return True
        else:
            print(
                f"アップロードに失敗しました。ステータスコード: {response.status_code}"
            )
            print(f"レスポンス: {response.text}")
            return False

    except Exception as e:
        print(f"ファイルアップロード中にエラーが発生しました: {e}")
        return False


def main():
    """メインの処理を実行"""
    # 現在の日時（タイムゾーン付き）
    now = datetime.now(ZoneInfo(TIMEZONE))

    # 期間の開始日と終了日を計算
    start_date, end_date = get_period_start_end(now)

    # JSONファイルのパスを設定
    json_filename = get_period_filename(start_date, end_date)
    json_file = os.path.join(DATA_DIR, json_filename)

    print(
        f"処理対象期間: {start_date.strftime('%Y/%m/%d')} ～ "
        f"{end_date.strftime('%Y/%m/%d')}"
    )
    print(f"対象ファイル: {json_file}")

    if not os.path.exists(json_file):
        print(f"メッセージファイルが見つかりません: {json_file}")
        return

    try:
        # JSONファイルを読み込む
        with open(json_file, "r", encoding="utf-8") as f:
            messages = json.load(f)

        # メッセージを処理
        processed_messages = {}
        for msg in messages:
            # タグと本文を分離
            text = msg.get(
                "text", msg.get("message", "")
            )  # textフィールドがない場合はmessageフィールドを使用
            tags, clean_text = extract_tags(text)
            tags_str = ", ".join(tags) if tags else ""  # '#'を含めたタグをそのまま使用

            # 時刻を分類
            received_at = msg.get("received_at", "")
            if received_at:
                timestamp = datetime.fromisoformat(received_at)
                time_str = timestamp.strftime("%H:%M")
                start_time, end_time, unknown_time = classify_time(clean_text, time_str)

                # ユーザー名でグループ化
                username = msg.get("username", "未設定")
                if username not in processed_messages:
                    processed_messages[username] = []

                # データを追加
                processed_messages[username].append(
                    {
                        "月": timestamp.strftime("%m"),
                        "日": timestamp.strftime("%d"),
                        "曜": get_weekday_jp(timestamp),
                        "出勤時刻": start_time,
                        "退社時刻": end_time,
                        "不明": unknown_time,
                        "タグ": tags_str,
                        "本文": clean_text,
                    }
                )

        # Excelファイルとして保存
        excel_filename = get_excel_filename(start_date, end_date)
        excel_file = os.path.join(DATA_DIR, excel_filename)

        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            # 各ユーザーのデータをシートとして保存
            for username, user_messages in processed_messages.items():
                # DataFrameを作成し、列の順序を指定
                df = pd.DataFrame(user_messages)
                df = df[
                    ["月", "日", "曜", "出勤時刻", "退社時刻", "不明", "タグ", "本文"]
                ]
                # シート名としてユーザー名を使用（31文字以内）
                sheet_name = username[:31]
                # シートとして保存（ヘッダーの太字を無効化）
                df.to_excel(
                    writer,
                    index=False,
                    sheet_name=sheet_name,
                )

                # ワークシートを取得
                worksheet = writer.sheets[sheet_name]

                # ヘッダーの太字を解除
                for cell in worksheet[1]:
                    cell.font = openpyxl.styles.Font(bold=False)

                # 罫線を削除
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = None

                # 月と日の列を数値書式に設定
                for row in worksheet.iter_rows(min_row=2):  # ヘッダー行をスキップ
                    # A列（月）の書式設定
                    row[0].number_format = "0"
                    # B列（日）の書式設定
                    row[1].number_format = "0"

                # 列幅の自動調整
                for idx, col in enumerate(df.columns):
                    # 列の最大文字数を計算（列名と内容の両方を考慮）
                    max_length = max(
                        df[col].astype(str).apply(len).max(),  # 内容の最大長
                        len(str(col)),  # 列名の長さ
                    )
                    # 文字幅から列幅を計算（1文字あたり1.2を掛けて余裕を持たせる）
                    adjusted_width = max_length * 1.2
                    # 列幅を設定（最小幅8、最大幅50）
                    worksheet.column_dimensions[chr(65 + idx)].width = min(
                        max(8, adjusted_width), 50
                    )

        print(f"Excelファイルを作成しました: {excel_file}")

        # WebDAVにアップロード
        if upload_to_webdav(excel_file, start_date, end_date):
            print("WebDAVへのアップロードが完了しました")
        else:
            print("WebDAVへのアップロードに失敗しました")

    except Exception as e:
        print(f"エラーが発生しました: {e}")
        return


if __name__ == "__main__":
    main()
